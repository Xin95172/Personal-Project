# =============================================================================
# build_path_file.py
# 台指選擇權 Path 檔案建構流程
#
# 輸入檔案：
#   1. 發行量加權股價指數.csv        — 台股加權指數（民國年格式）
#   2. 2025台灣指數選擇權.csv        — 2025 年選擇權到期日（xlsx 格式）
#   3. 202512-202601台指選擇權.csv   — 補充到期日（xlsx 格式，補跨年缺口）
#
# 輸出檔案：
#   Path_發行量加權指數_完整_Rf_final.xlsx
#     欄位：Date | File | S0 | Contract | Contract Expiry Date | Maturity | Rf
#
# 執行環境：Python 3.8+，需安裝 pandas、openpyxl
#   pip install pandas openpyxl
#
# Google Colab 使用說明：
#   1. 上傳本程式及三個輸入檔至 Colab
#   2. 取消 Step 0 中 files.upload() 的註解
#   3. 執行所有儲存格
# =============================================================================

import pandas as pd
import urllib.request
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── 檔案路徑設定 ─────────────────────────────────────────────────────────────
INDEX_FILE   = '發行量加權股價指數.csv'         # 台股加權指數（民國年）
OPTIONS_FILE = '2025台灣指數選擇權.csv'         # 選擇權到期日主檔（xlsx）
SUPP_FILE    = '202512-202601台指選擇權.csv'    # 跨年補充到期日（xlsx）
OUTPUT_FILE  = 'Path_發行量加權指數_完整_Rf_final.xlsx'

# ── Colab 上傳（本地執行請保持註解）──────────────────────────────────────────
# from google.colab import files
# uploaded = files.upload()   # 一次選擇以上三個輸入檔


# =============================================================================
# Step 1：讀取台股加權指數，轉換民國年 → 西元年，生成 File 欄位
# =============================================================================
def roc_to_western(roc_date: str):
    """將民國年日期字串（e.g. '114/01/02'）轉為 (西元日期字串, 年, 月, 日)"""
    parts = roc_date.split('/')
    year  = int(parts[0]) + 1911
    month = int(parts[1])
    day   = int(parts[2])
    return f"{year}/{month}/{day}", year, month, day


print("=" * 60)
print("Step 1：處理台股加權指數資料")
print("=" * 60)

idx_df = pd.read_csv(INDEX_FILE, encoding='utf-8-sig')
idx_df.columns = ['Date_ROC', 'S0']

parsed          = idx_df['Date_ROC'].apply(roc_to_western)
idx_df['Date']  = parsed.apply(lambda x: x[0])
idx_df['_year'] = parsed.apply(lambda x: x[1])
idx_df['_mon']  = parsed.apply(lambda x: x[2])
idx_df['_day']  = parsed.apply(lambda x: x[3])

# File 欄位：OptionsDaily_YYYY_MM_DD.csv（月份、日期補零）
idx_df['File'] = idx_df.apply(
    lambda r: f"OptionsDaily_{r['_year']}_{r['_mon']:02d}_{r['_day']:02d}.csv",
    axis=1
)

# S0：移除千分位逗號，轉 float
idx_df['S0'] = idx_df['S0'].str.replace(',', '').astype(float)

path = idx_df[['Date', 'File', 'S0']].copy()
path['Date'] = pd.to_datetime(path['Date'])

print(f"  交易日共 {len(path)} 筆：{path['Date'].min().date()} ~ {path['Date'].max().date()}")


# =============================================================================
# Step 2：讀取選擇權到期日，篩選月選擇權
#   月選擇權定義：契約月份不含 'W'
# =============================================================================
def load_monthly_options(filepath: str) -> pd.DataFrame:
    """
    讀取選擇權到期日 xlsx，回傳月選擇權 DataFrame
    欄位：最後結算日 (datetime)、契約月份 (str)
    """
    raw = pd.read_excel(filepath, header=None)
    # 找真正的表頭列（含「最後結算日」）
    header_idx = raw[
        raw.apply(lambda r: r.astype(str).str.contains('最後結算日').any(), axis=1)
    ].index[0]
    raw.columns = raw.iloc[header_idx]
    raw = raw.iloc[header_idx + 1:].dropna(how='all').reset_index(drop=True)

    col_expiry   = [c for c in raw.columns if '結算日' in str(c)][0]
    col_contract = [c for c in raw.columns if '契約' in str(c) or '期約' in str(c)][0]

    opt = raw[[col_expiry, col_contract]].copy()
    opt.columns = ['最後結算日', '契約月份']
    opt['最後結算日'] = pd.to_datetime(opt['最後結算日'])
    opt['契約月份']   = opt['契約月份'].astype(str).str.strip()

    monthly = opt[~opt['契約月份'].str.contains('W')].copy()
    return monthly.sort_values('最後結算日').reset_index(drop=True)


print("\n" + "=" * 60)
print("Step 2：讀取月選擇權到期日")
print("=" * 60)

monthly_main = load_monthly_options(OPTIONS_FILE)
monthly_supp = load_monthly_options(SUPP_FILE)

# 合併（去除重複，以到期日為 key）
monthly_all = (
    pd.concat([monthly_main, monthly_supp], ignore_index=True)
    .drop_duplicates(subset=['最後結算日'])
    .sort_values('最後結算日')
    .reset_index(drop=True)
)

print(f"  月選擇權共 {len(monthly_all)} 筆：")
print(monthly_all.to_string(index=False))


# =============================================================================
# Step 3：對每個交易日找「最近月選擇權」
#   條件：最後結算日 - 交易日 >= 1 日曆天
#   結果：Contract、Contract Expiry Date、Maturity（日曆天數）
# =============================================================================
def find_nearest_monthly(trade_date: pd.Timestamp, monthly_df: pd.DataFrame):
    """
    找 trade_date 之後距離最近的月選擇權
    條件：(最後結算日 - trade_date).days >= 1
    回傳：(契約月份, 到期日字串, 剩餘天數) 或 (None, None, None)
    """
    candidates = monthly_df[
        (monthly_df['最後結算日'] - trade_date).dt.days >= 1
    ]
    if candidates.empty:
        return None, None, None
    nearest  = candidates.loc[candidates['最後結算日'].idxmin()]
    maturity = (nearest['最後結算日'] - trade_date).days
    return nearest['契約月份'], str(nearest['最後結算日'].date()), maturity


print("\n" + "=" * 60)
print("Step 3：配對月選擇權合約")
print("=" * 60)

results = path['Date'].apply(
    lambda d: pd.Series(
        find_nearest_monthly(d, monthly_all),
        index=['Contract', 'Contract Expiry Date', 'Maturity']
    )
)
path = pd.concat([path, results], axis=1)

missing_count = path['Contract'].isna().sum()
print(f"  配對完成，缺失筆數：{missing_count}")

# 驗證合約切換點
changes = path[path['Contract'] != path['Contract'].shift()]
print("\n  合約切換點：")
print(changes[['Date', 'Contract', 'Contract Expiry Date', 'Maturity']]
      .assign(Date=lambda x: x['Date'].dt.strftime('%Y/%m/%d'))
      .to_string(index=False))


# =============================================================================
# Step 4：抓取台銀一年期定儲機動利率（Rf）
#   來源：https://rate.bot.com.tw/twd/csv/YYYY-MM-DD
#   欄位：一年~未滿二年 一般定期儲蓄存款 機動利率
#   以「實施日期」快取，避免重複下載
# =============================================================================
def fetch_bot_rate(date_str: str):
    """
    從台銀抓取指定日期的一年期定儲機動利率
    date_str : 'YYYY-MM-DD'
    回傳     : (impl_date: str, rate: float)  rate 為年利率小數
    """
    url = f'https://rate.bot.com.tw/twd/csv/{date_str}'
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=15) as resp:
        raw = resp.read().decode('utf-8-sig')

    lines   = raw.strip().split('\n')
    impl_dt = lines[0].split(',')[0].replace('實施日期：', '').strip()
    headers = [h.strip() for h in lines[1].split(',')]
    values  = [v.strip() for v in lines[2].split(',')]

    idx = next(
        i for i, h in enumerate(headers)
        if '一年' in h and '未滿' in h and '一般定期儲蓄' in h and '機動' in h
    )
    return impl_dt, float(values[idx]) / 100


print("\n" + "=" * 60)
print("Step 4：抓取台銀 Rf 利率")
print("=" * 60)
print(f"  共 {len(path)} 個交易日...")

cache, rf_list = {}, []

for i, row in path.iterrows():
    date_str = row['Date'].strftime('%Y-%m-%d')
    try:
        impl_dt, rate = fetch_bot_rate(date_str)
        if impl_dt not in cache:
            cache[impl_dt] = rate
            print(f"  [{i+1:3d}/{len(path)}] {date_str}  實施日期={impl_dt}  Rf={rate*100:.4f}%  ← 新利率")
        else:
            rate = cache[impl_dt]
        rf_list.append(rate)
    except Exception as e:
        print(f"  [{i+1:3d}] {date_str} 失敗：{e}，沿用上筆")
        rf_list.append(rf_list[-1] if rf_list else None)
    time.sleep(0.3)

path['Rf'] = rf_list
print(f"\n  Rf 唯一值：{sorted(set(rf_list))}")


# =============================================================================
# Step 5：整理欄位，輸出 Excel（置中對齊）
# =============================================================================
print("\n" + "=" * 60)
print("Step 5：輸出 Excel")
print("=" * 60)

# 日期格式、型態整理
path['Date']     = path['Date'].dt.strftime('%Y/%-m/%-d')
path['Maturity'] = path['Maturity'].astype('Int64')   # nullable int（允許 NaN）

col_order = ['Date', 'File', 'S0', 'Contract', 'Contract Expiry Date', 'Maturity', 'Rf']
path = path[col_order]

# 寫入 xlsx
path.to_excel(OUTPUT_FILE, index=False, engine='openpyxl')

# ── 套用樣式 ──────────────────────────────────────────────────────────────────
wb = load_workbook(OUTPUT_FILE)
ws = wb.active

header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF')
center      = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 表頭
for cell in ws[1]:
    cell.alignment = center
    cell.font      = header_font
    cell.fill      = header_fill
    cell.border    = thin_border

# 資料列
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = center
        cell.border    = thin_border

# 自動欄寬
for col in ws.columns:
    max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

# 凍結首列
ws.freeze_panes = 'A2'
wb.save(OUTPUT_FILE)

print(f"  ✓ 已儲存：{OUTPUT_FILE}")
print(f"  總筆數：{len(path)}")
print(f"\n=== 最終結果預覽（前5筆）===")
print(path.head().to_string(index=False))
print(f"\n=== 最終結果預覽（後5筆）===")
print(path.tail().to_string(index=False))

# ── Colab 下載（本地執行保持註解）────────────────────────────────────────────
# from google.colab import files
# files.download(OUTPUT_FILE)
