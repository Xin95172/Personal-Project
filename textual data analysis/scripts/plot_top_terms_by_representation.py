from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(r"D:\Github\Personal-Project\textual data analysis")
WORKSPACE_ROOT = Path(r"C:\Users\ownme\Documents\New project")
REPORTS_ROOT = PROJECT_ROOT / "artifacts" / "reports"
LEXICON_ROOT = PROJECT_ROOT / "resources" / "lexicons"
OUTPUT_ROOT = WORKSPACE_ROOT / "outputs" / "top_terms_by_representation"

sys.path.insert(0, str(PROJECT_ROOT))
from utils.dtm_utils import custom_tokenizer  # noqa: E402


DATASETS = {
    "administrative_win_lose_mixed": "Administrative",
    "civil_win_lose_mixed": "Civil",
    "criminal_win_lose_mixed": "Criminal",
    "cwc_win_lose_mixed": "Criminal with Attached Civil",
}

REPRESENTATIONS = {
    "bow": {
        "label": "BoW",
        "file_stem": "bow",
        "metric": "relative_frequency",
        "x_label": "Relative word frequency",
    },
    "tf": {
        "label": "TF",
        "file_stem": "tf",
        "metric": "mean_tf",
        "x_label": "Mean TF weight",
    },
    "tfidf": {
        "label": "TF-IDF",
        "file_stem": "tfidf",
        "metric": "mean_tfidf",
        "x_label": "Mean TF-IDF weight",
    },
}

COLORS = {
    "Administrative": "#ef6f61",
    "Civil": "#2f9e44",
    "Criminal": "#00a6b4",
    "Criminal with Attached Civil": "#9b5de5",
}

GENERIC_LEGAL_TERMS = {
    "原告",
    "被告",
    "上訴人",
    "被上訴人",
    "參加人",
    "申請人",
    "相對人",
    "抗告人",
    "告訴人",
    "被告人",
    "訴願人",
    "代理人",
    "代表人",
    "訴訟",
    "法院",
    "法官",
    "判決",
    "裁定",
    "上訴",
    "聲請",
    "申請",
    "提起",
    "起訴",
    "審理",
    "審查",
    "辯論",
    "程序",
    "請求",
    "主張",
    "追加",
    "記載",
    "管轄",
    "主文",
    "理由",
    "事實",
    "證據",
    "規定",
    "方面",
    "事件",
    "案件",
    "法律",
    "法條",
    "民事",
    "刑事",
    "行政",
    "我國",
    "系爭",
    "決定",
    "概要",
    "不服",
    "公司",
    "行為",
    "附帶",
    "民國",
    "年度",
    "審定",
    "處分",
    "駁回",
    "核駁",
    "變更",
    "事項",
    "聲明",
    "第一",
    "第二",
    "第三",
    "附表",
    "附件",
    "簡易",
    "檢察官",
    "經濟部",
    "智慧財產局",
    "智財局",
    "智慧",
    "財產",
    "財產權",
    "管轄權",
    "管轄",
    "涉外",
    "國際",
    "明文",
    "准許",
    "准予",
    "撤銷",
    "原處分",
    "訴願",
    "訴願決定",
    "公告",
    "編為",
    "定有",
    "引用",
    "處刑",
    "有限公司",
    "裁判",
    "任何",
    "計算",
    "智專",
    "指定",
    "施行",
    "有違",
    "組織法",
    "審理法",
    "補充",
    "權人",
    "陳述",
    "行為地",
    "起訴狀",
    "意圖",
    "參加",
    "附圖",
    "修正",
    "成立",
    "給付",
    "股份",
    "書狀",
    "提出",
    "中華民國",
    "明知",
    "更正為",
    "公分",
    "發生",
    "言詞",
    "終結",
    "費用",
    "負擔",
    "假執行",
    "供擔保",
    "送達",
    "本院",
    "本件",
    "前揭",
    "上開",
    "下稱",
    "理由",
    "部分",
}


def read_lines(path: Path) -> list[str]:
    return path.read_text(encoding="utf-8").splitlines()


def normalize_token(token: str) -> str:
    return unicodedata.normalize("NFKC", str(token)).lower().strip()


def strip_phrases(series: pd.Series, phrases: list[str]) -> pd.Series:
    texts = series.astype(str).map(normalize_token)
    if not phrases:
        return texts
    for phrase in phrases:
        phrase_norm = normalize_token(phrase)
        if not phrase_norm or " " not in phrase_norm:
            continue
        pattern = re.compile(re.escape(phrase_norm))
        texts = texts.apply(lambda value: pattern.sub(" ", value))
    return texts


def make_analyzer(delete_words: list[str], phrases: list[str]):
    banned = {normalize_token(term) for term in delete_words if term and " " not in str(term)}
    phrase_patterns = [
        re.compile(re.escape(normalize_token(phrase)))
        for phrase in phrases
        if isinstance(phrase, str) and " " in phrase and normalize_token(phrase)
    ]

    def analyzer(doc: str) -> list[str]:
        if phrase_patterns:
            doc = normalize_token(doc)
            for pattern in phrase_patterns:
                doc = pattern.sub(" ", doc)
        tokens = [normalize_token(token) for token in custom_tokenizer(doc)]
        return [token for token in tokens if token and token not in banned]

    return analyzer


def add_generic_legal_terms(delete_words: list[str]) -> list[str]:
    merged = list(delete_words)
    merged.extend(GENERIC_LEGAL_TERMS)
    return merged


def iter_word_seg_texts(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        try:
            text_col = header.index("Word Segmentation")
        except ValueError as exc:
            raise KeyError(f"'Word Segmentation' column not found in {path}") from exc
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            value = row[text_col]
            if value:
                yield str(value)
    finally:
        workbook.close()


def load_doc_counters(path: Path, analyzer) -> tuple[list[Counter], Counter, Counter]:
    doc_counters: list[Counter] = []
    total_counts: Counter = Counter()
    doc_counts: Counter = Counter()
    for text in iter_word_seg_texts(path):
        counter = Counter(analyzer(text))
        if not counter:
            continue
        doc_counters.append(counter)
        total_counts.update(counter)
        doc_counts.update(counter.keys())
    return doc_counters, total_counts, doc_counts


def allowed_terms(doc_counts: Counter, n_docs: int) -> set[str]:
    max_df = int(np.floor(0.98 * n_docs))
    return {term for term, count in doc_counts.items() if count >= 2 and count <= max_df}


def score_bow(total_counts: Counter, vocab: set[str]) -> dict[str, float]:
    scores = {term: float(count) for term, count in total_counts.items() if term in vocab}
    denominator = sum(scores.values())
    if denominator:
        scores = {term: score / denominator for term, score in scores.items()}
    return scores


def score_tf(doc_counters: list[Counter], vocab: set[str]) -> dict[str, float]:
    accum: Counter = Counter()
    for counter in doc_counters:
        filtered = {term: count for term, count in counter.items() if term in vocab}
        norm = np.sqrt(sum(count * count for count in filtered.values()))
        if not norm:
            continue
        for term, count in filtered.items():
            accum[term] += count / norm
    return {term: value / len(doc_counters) for term, value in accum.items()}


def score_tfidf(doc_counters: list[Counter], doc_counts: Counter, vocab: set[str]) -> dict[str, float]:
    n_docs = len(doc_counters)
    idf = {term: np.log((1 + n_docs) / (1 + doc_counts[term])) + 1 for term in vocab}
    accum: Counter = Counter()
    for counter in doc_counters:
        weighted = {term: count * idf[term] for term, count in counter.items() if term in vocab}
        norm = np.sqrt(sum(value * value for value in weighted.values()))
        if not norm:
            continue
        for term, value in weighted.items():
            accum[term] += value / norm
    return {term: value / n_docs for term, value in accum.items()}


def compute_top_terms_for_all_representations(
    dataset_name: str,
    dataset_label: str,
    leakage_variant: str,
    analyzer,
    top_n: int,
) -> dict[str, pd.DataFrame]:
    word_seg_path = REPORTS_ROOT / dataset_name / leakage_variant / "word_seg.xlsx"
    print(f"Reading {word_seg_path}", flush=True)
    doc_counters, total_counts, doc_counts = load_doc_counters(word_seg_path, analyzer)
    vocab = allowed_terms(doc_counts, len(doc_counters))
    print(f"  docs={len(doc_counters):,}, vocab={len(vocab):,}", flush=True)

    all_scores = {
        "bow": score_bow(total_counts, vocab),
        "tf": score_tf(doc_counters, vocab),
        "tfidf": score_tfidf(doc_counters, doc_counts, vocab),
    }
    frames = {}
    for representation, scores in all_scores.items():
        metric = REPRESENTATIONS[representation]["metric"]
        top_items = sorted(scores.items(), key=lambda item: item[1], reverse=True)[:top_n]
        frames[representation] = pd.DataFrame(
            {
                "dataset": dataset_name,
                "dataset_label": dataset_label,
                "leakage_variant": leakage_variant,
                "representation": REPRESENTATIONS[representation]["label"],
                "rank": np.arange(1, len(top_items) + 1),
                "term": [term for term, _ in top_items],
                metric: [score for _, score in top_items],
            }
        )
    return frames


def configure_matplotlib() -> None:
    plt.rcParams["font.sans-serif"] = [
        "Microsoft JhengHei",
        "Microsoft YaHei",
        "SimHei",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["figure.dpi"] = 130


def plot_facets(data: pd.DataFrame, representation: str, leakage_variant: str, top_n: int, suffix: str) -> None:
    rep = REPRESENTATIONS[representation]
    metric = rep["metric"]
    fig, axes = plt.subplots(2, 2, figsize=(12, 9), constrained_layout=True)
    axes = axes.ravel()

    for ax, dataset_label in zip(axes, DATASETS.values()):
        subset = data[data["dataset_label"] == dataset_label].sort_values(metric, ascending=True)
        ax.barh(
            subset["term"],
            subset[metric],
            color=COLORS[dataset_label],
            edgecolor="none",
        )
        ax.set_title(dataset_label, fontsize=11, pad=8, color="white", backgroundcolor="#9ca3af")
        ax.grid(axis="x", color="#d1d5db", linewidth=0.7)
        ax.grid(axis="y", color="#e5e7eb", linewidth=0.5)
        ax.set_axisbelow(True)
        ax.tick_params(axis="both", labelsize=9)
        for spine in ax.spines.values():
            spine.set_color("#d1d5db")

    fig.supxlabel(rep["x_label"], fontsize=11, fontweight="bold")
    fig.suptitle(
        f"Top {top_n} {rep['label']} Terms by Case Type ({leakage_variant}{suffix.replace('_', ' ')})",
        fontsize=14,
        fontweight="bold",
    )

    out_dir = OUTPUT_ROOT / leakage_variant
    out_dir.mkdir(parents=True, exist_ok=True)
    png_path = out_dir / f"top_terms_by_case_type_{rep['file_stem']}_{leakage_variant}{suffix}.png"
    pdf_path = out_dir / f"top_terms_by_case_type_{rep['file_stem']}_{leakage_variant}{suffix}.pdf"
    csv_path = out_dir / f"top_terms_by_case_type_{rep['file_stem']}_{leakage_variant}{suffix}.csv"

    data.to_csv(csv_path, index=False, encoding="utf-8-sig")
    fig.savefig(png_path, bbox_inches="tight", dpi=300)
    fig.savefig(pdf_path, bbox_inches="tight")
    plt.close(fig)
    print(png_path)
    print(pdf_path)
    print(csv_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--leakage-variant", choices=["no_leakage", "with_leakage"], default="no_leakage")
    parser.add_argument("--top-n", type=int, default=15)
    parser.add_argument(
        "--strip-phrases",
        action="store_true",
        help="Also remove multi-token stop/delete phrases before vectorization. This is slower.",
    )
    parser.add_argument(
        "--exclude-generic-legal-terms",
        action="store_true",
        help="Remove common party labels and procedural/legal boilerplate terms from the top-term charts.",
    )
    args = parser.parse_args()

    configure_matplotlib()
    delete_words = read_lines(LEXICON_ROOT / "delete_vocab.txt")
    suffix = "_content_terms" if args.exclude_generic_legal_terms else ""
    if args.exclude_generic_legal_terms:
        delete_words = add_generic_legal_terms(delete_words)
    stop_words = read_lines(LEXICON_ROOT / "stopwords-ch-jiebar-zht.txt") if args.strip_phrases else []
    phrases = (
        [
            word
            for word_list in (stop_words, delete_words)
            for word in word_list
            if isinstance(word, str) and " " in word
        ]
        if args.strip_phrases
        else []
    )
    analyzer = make_analyzer(delete_words, phrases)

    rows_by_representation = {representation: [] for representation in REPRESENTATIONS}
    for dataset_name, dataset_label in DATASETS.items():
        frames = compute_top_terms_for_all_representations(
            dataset_name=dataset_name,
            dataset_label=dataset_label,
            leakage_variant=args.leakage_variant,
            analyzer=analyzer,
            top_n=args.top_n,
        )
        for representation, frame in frames.items():
            rows_by_representation[representation].append(frame)

    for representation, rows in rows_by_representation.items():
        plot_facets(pd.concat(rows, ignore_index=True), representation, args.leakage_variant, args.top_n, suffix)


if __name__ == "__main__":
    main()
